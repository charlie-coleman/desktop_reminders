import tkinter as tk
from tkinter import messagebox
from tkinter import filedialog
import datetime
import openpyxl as xl
import time

setup = True
tk.Tk().withdraw()
with open('.\\config.txt','r') as f:
    lines = f.readlines()
for line in lines:
    if 'Workbook_File_Path: ' in line:
        setup = False
        workbook_file_path = line[len('Workbook_File_Path: '):]
if setup:
    f = open('.\\config.txt', 'w')
    workbook_file_path = filedialog.askopenfilename()
    f.writelines('Workbook_File_Path: ' + workbook_file_path)
    f.close()

while True:
    print("Triggered!")
    wb = xl.load_workbook(workbook_file_path)
    ws = wb.active
    current_dt = datetime.datetime.now().replace(second=0, microsecond=0)
    time_arr = []
    max_num = ws.cell('E1').value
    for i in range(1, max_num+1):
        time_arr.append(ws.cell(row=i, column=1).value)
    if None in time_arr:
        first_open = time_arr.index(None)+1
    else:
        first_open = max_num+1
    for i in range(len(time_arr)):
        time_val = time_arr[i]
        if time_val is not None and datetime.datetime.strptime(time_val, '%Y-%m-%d %H:%M:%S') < current_dt:
            message_out = ws.cell(row=i+1, column=2).value
            messagebox.showinfo(title='Reminder', message=message_out)
            ws.cell(row=i+1, column=1).value = None
            ws.cell(row=i+1, column=2).value = None
            time_arr.pop(i)
        time_arr = []
        for j in range(1, max_num+1):
            time_arr.append(ws.cell(row=j, column=1).value)
    for i in reversed(range(len(time_arr))):
        time_val = time_arr[i]
        if time_val is None:
            ws.cell('E1').value -= 1
            max_num -= 1
            time_arr.pop(i)
        else:
            break
    wb.save(workbook_file_path)
    time.sleep(50)
