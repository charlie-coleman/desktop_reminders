import sys
import getopt
import datetime
import openpyxl as xl
from tkinter import filedialog


def main(argv):
    setup = True
    lines = []
    try:
        with open('c:\\PythonScripts\\desktop_reminders\\config.txt', 'r') as f:
            lines = f.readlines()
    except FileNotFoundError:
        setup = True
    for line in lines:
        if 'Workbook_File_Path: ' in line:
            setup = False
            workbook_file_path = line[len('Workbook_File_Path: '):]
    if setup:
        f = open('c:\\PythonScripts\\desktop_reminders\\config.txt', 'w')
        workbook_file_path = filedialog.askopenfilename()
        f.writelines('Workbook_File_Path: ' + workbook_file_path)
        f.close()
    wb = xl.load_workbook(workbook_file_path)
    ws = wb.active
    time_arr = []
    max_num = ws.cell('E1').value
    for i in range(1, max_num+1):
        time_arr.append(ws.cell(row=i, column=1).value)
    if None in time_arr:
        first_open = time_arr.index(None)+1
    else:
        first_open = max_num+1
    date = ''
    time = ''
    message = ''
    try:
        opts, args = getopt.getopt(argv, 'hld:t:m:r:', ['help', 'list', 'date=', 'time=', 'message=', 'remove='])
    except getopt.GetoptError:
        print('add_reminder -d <date (YYYY-mm-dd)> -t <time (HH:MM)> -m <message>')
        sys.exit(2)
    add = True
    for opt, arg in opts:
        if opt in ('-h', '--help'):
            print('add_reminder -d <date (YYYY-mm-dd)> -t <time (HH:MM)> -m <message>')
            sys.exit()
        elif opt in ('-l', '--list'):
            for i in range(len(time_arr)):
                if time_arr[i] is not None:
                    print('Index: %s\tTime: %s\tMessage: %s' % (i+1, time_arr[i], ws.cell(row=i+1, column=2).value))
            sys.exit()
        elif opt in ('-d', '--date'):
            date = arg
        elif opt in ('-t', '--time'):
            time = arg
        elif opt in ('-m', '--message'):
            message = arg
        elif opt in ('-r', '--remove'):
            ws.cell(row=int(arg), column=1).value = None
            ws.cell(row=int(arg), column=2).value = None
            wb.save(workbook_file_path)
            sys.exit()
    if date == '':
        date = str(datetime.date.today())
    if time == '':
        time = datetime.datetime.now() + datetime.timedelta(minutes=15)
        time = time.strftime('%H:%M')
        time = str(time)
    if message == '':
        message = 'Reminder'
    dt = datetime.datetime.strptime(date + ' ' + time, '%Y-%m-%d %H:%M')
    ws.cell(row=first_open, column=1).value = str(dt)
    ws.cell(row=first_open, column=2).value = message
    print('Time: %s\nMessage: %s' % (str(dt), message))
    if first_open > max_num:
        max_num = first_open
        ws.cell('e1').value = max_num
    wb.save(workbook_file_path)

if __name__ == '__main__':
    main(sys.argv[1:])